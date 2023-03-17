import openpyxl

reports = "C:\\Users\\60035675\\Desktop\\PythonCourse\\eHealth\\Final-Delta-CHW\\OldAndNewReport.xlsx"

wb = openpyxl.load_workbook(reports)
sheet = wb.active

clean_wb = openpyxl.Workbook()
clean_sheet = clean_wb.active

clean_sheet.append(['AccessionNumber', 'ReportText', 'ORUReportText'])


# Loop through each row in the sheet
for row in sheet.iter_rows(min_row=2, values_only=True):
    report_text = row[1].strip()  
    oru_report_text = row[2].replace('\\X0D', '\n').replace('\\X0A', '').replace('\\', '').strip()  # Remove special characters and strip white spaces from column C
    report_text = report_text.replace('\n','')
    oru_report_text = report_text.replace('\n','')

    # Write cleaned data to the new sheet
    clean_sheet.append([row[0], report_text, oru_report_text])

    # Check if the value in columns B and C are different
    if report_text != oru_report_text:
        print(f"Row {row[0]} has different values in columns B and C")

clean_wb.save('cleaned_data.xlsx')